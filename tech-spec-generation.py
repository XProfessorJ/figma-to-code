import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font

# Constants
HTML_FILE = 'CBOL-template.html'
EXCEL_FILE = 'output.xlsx'
FUNCTION_NAME = 'TradeNow'
CLASS_PREFIXES = {
    'Title': 'Hdr',
    'Button': 'Btn',
    'Label': 'Lbl',
    'ListItem': 'Txt',
    'Value': None
}

def read_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        return file.read()

def extract_elements_with_content(html_content):
    soup = BeautifulSoup(html_content, 'html.parser')
    return soup.find_all(string=True)

def generate_label_id(class_name, content):
    prefix = next((prefix for key, prefix in CLASS_PREFIXES.items() if key in class_name), 'Txt')
    if prefix:
        return f"{prefix}_{FUNCTION_NAME}_{content.replace(' ', '_')}"
    elif 'Value' in class_name:
        return f"Lbl_{FUNCTION_NAME}_{content.replace(' ', '_')}_Placeholder"
    return 'NA'

def process_elements(elements):
    data = []
    for element in elements:
        parent = element.parent
        if parent and parent.get('class') and element.strip():
            class_name = ' '.join(parent.get('class'))
            content = element.strip()
            label_id = generate_label_id(class_name, content)
            cta = "Yes" if 'Button' in class_name else 'NA'
            content_biz_managed = "Yes" if label_id != 'NA' else 'NA'

            data.append({
                'Screen Element': content,
                'Label ID': label_id,
                'CTA': cta,
                'Content-Biz Managed': content_biz_managed,
                'DataType(input)': 'NA',
                'Display If': None,
                'Country': None,
                'Channel': None,
                'API(For API\'s refer API sheet)': None
            })

            if "Value" in class_name:
                data.append({
                    'Screen Element': '{{Value}}',
                    'Label ID': 'NA',
                    'CTA': 'NA',
                    'Content-Biz Managed': 'NA',
                    'DataType(input)': "NA",
                    'Display If': None,
                    'Country': None,
                    'Channel': None,
                    'API(For API\'s refer API sheet)': None
                })
    return data

def write_to_excel(data, file_path):
    module_info = [
        {'Module Name': 'MODULE', 'Module Description': 'eBrokerage'},
        {'Module Name': 'SUB MODULE', 'Module Description': 'eBrokerge Buy'},
        {'Module Name': 'PAGENAME', 'Module Description': 'TradeNow_Buy_100'}
    ]
    df_description = pd.DataFrame(module_info)
    df = pd.DataFrame(data)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_description.to_excel(writer, index=False, header=False)
        df.to_excel(writer, index=False, startrow=3)

    wb = load_workbook(file_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.font = Font(bold=True)
    wb.save(file_path)

def main():
    html_content = read_html(HTML_FILE)
    elements = extract_elements_with_content(html_content)
    data = process_elements(elements)
    write_to_excel(data, EXCEL_FILE)
    loaded_df = pd.read_excel(EXCEL_FILE)
    print(loaded_df)

if __name__ == "__main__":
    main()